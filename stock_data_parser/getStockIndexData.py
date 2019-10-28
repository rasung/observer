import FinanceDataReader as fdr
from openpyxl import Workbook
import os
import sys
import pandas_datareader as pdr

def getFinanceData(index_symbol, start_year, file_name):
	write_wb = Workbook()
	#Sheet1에다 입력
	write_ws = write_wb.active

	df = fdr.DataReader(index_symbol, start_year)
	for i in range(len(df)):
			write_ws.cell(i+1, 1, df['Close'].index[i].date())
			write_ws.cell(i+1, 2, df['Close'].values[i])

	write_wb.save(os.getcwd() + '/' + file_name)


def getStockData():
	#한국
	#getFinanceData('KS11', '2007', 'Chart_KS11.xlsx')
	#getFinanceData('KQ11', '1990', 'Chart_KQ11.xlsx')

	#미국
	getFinanceData('DJI', '1990', 'Chart_DJI.xlsx')
	getFinanceData('IXIC', '2009', 'Chart_IXIC.xlsx')

	#중국
	#getFinanceData('FTXIN9', '1990', 'Chart_FTXIN9.xlsx')
	
	#홍콩
	#getFinanceData('HSI', '1990', 'Chart_HSI.xlsx')

	#일본
	#getFinanceData('JP225', '1990', 'Chart_JP225.xlsx')


def getExchangeRateData():
	#달러당 원화 환율
	getFinanceData('USD/KRW', '2009', 'Chart_USD_KRW.xlsx')

	#달러당 유로화 환율
	getFinanceData('USD/EUR', '2009', 'Chart_USD_EUR.xlsx')

	#달러당 엔화 환율
	getFinanceData('USD/JPY', '2009', 'Chart_USD_JPY.xlsx')


def getGoldData(start_year, file_name):
	write_wb = Workbook()
	#Sheet1에다 입력
	write_ws = write_wb.active

	df_gold = pdr.DataReader('GOLDAMGBD228NLBM', 'fred', start=start_year)
	for i in range(len(df_gold)):
		write_ws.cell(i+1, 1 , df_gold['GOLDAMGBD228NLBM'].index[i].date())
		write_ws.cell(i+1, 2, df_gold['GOLDAMGBD228NLBM'].values[i])
	
	write_wb.save(os.getcwd() + '/' + file_name)


def getOilData(start_year, file_name):
	write_wb = Workbook()
	#Sheet1에다 입력
	write_ws = write_wb.active

	df_gold = pdr.DataReader(['POILDUBUSDM', 'POILWTIUSDM', 'POILBREUSDM'], 'fred', start=start_year)
	for i in range(len(df_gold)):
		write_ws.cell(i+1, 1 , df_gold['POILDUBUSDM'].index[i].date())
		write_ws.cell(i+1, 2, df_gold['POILDUBUSDM'].values[i])
		write_ws.cell(i+1, 3, df_gold['POILWTIUSDM'].values[i])
		write_ws.cell(i+1, 4, df_gold['POILBREUSDM'].values[i])
	
	write_wb.save(os.getcwd() + '/' + file_name)


def main():
	getStockData()
	#getExchangeRateData()
	#getGoldData('1990-01-01', 'GOLD.xlsx')
	#getOilData('1990-01-01', 'OIL.xlsx')


if __name__ == "__main__":
	main()
